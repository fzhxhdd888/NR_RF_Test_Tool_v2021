#!/usr/bin/env python 
# -*- coding: utf-8 -*-
# @Time    : 2020/11/26 17:01
# @Author  : Feng Zhaohui
# @Site    : 
# @File    : reporthandle.py
# @Software: PyCharm

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import os
import time
import global_elements

# from docx import Document
# from docx.shared import Inches
# import shutil
# from win32com.client import Dispatch
# import win32api
# import win32con


font_pass = Font(color="008000")
font_fail = Font(color="FF0000")
font_inconclusive = Font(color="A52A2A")


# Judgement判断
def judgement_handle(result, lowlimit, highlimit):
    if result != '-999':
        try:
            if lowlimit == 'None' and highlimit == 'None':
                judgement = 'Passed'
            elif lowlimit == 'None' and highlimit != 'None':
                if float(highlimit) - float(result) >= 0:
                    judgement = 'Passed'
                else:
                    judgement = 'Failed'
            elif lowlimit != 'None' and highlimit == 'None':
                if float(result) - float(lowlimit) >= 0:
                    judgement = 'Passed'
                else:
                    judgement = 'Failed'
            else:
                if float(result) - float(lowlimit) >= 0 and float(highlimit) - float(result) >= 0:
                    judgement = 'Passed'
                else:
                    judgement = 'Failed'
        except:
            judgement = 'Inconclusive'
    else:
        judgement = 'Inconclusive'

    return judgement


# NR 信道号与频率的换算
def NreftoFre(Nref):
    if 0 <= int(Nref) <= 599999:
        Fref = str(0 + (5 * (int(Nref) - 0)) / 1000)
    elif 600000 <= int(Nref) <= 2016666:
        Fref = str(3000 + (15 * (int(Nref) - 600000)) / 1000)
    else:
        Fref = '-'
    return Fref


# 测试得到的数据输出到报告和界面
def Reporttool(result, lowlimit, highlimit, unit):
    try:
        if global_elements.NR_type_r == 'NR_SA':
            ReporttoolforSA(result, lowlimit, highlimit, unit)
        elif global_elements.NR_type_r == 'NR_NSA':
            ReporttoolforNSA(result, lowlimit, highlimit, unit)
    except Exception as e:
        global_elements.emitsingle.thread_exitSingle.emit('Creat excel report Faild, Error: ' + e.__doc__)


# 测试得到的数据输出到报告和界面
def ReporttoolforSA(result, lowlimit, highlimit, unit):
    if global_elements.isStatusError == False:  # 判断是否已经终止进程
        if not os.access(global_elements.REPORTPATH, os.F_OK):  # 先判断report文件是否存存在， 不存在就新建
            reportwb = Workbook()
            sheets_list = reportwb.sheetnames
            if len(sheets_list) > 1:
                for i in range(1, len(sheets_list)):
                    reportwb.remove(reportwb[sheets_list[i]])  # 只保存一个sheet

            reportwb.save(global_elements.REPORTPATH)
            reportwb.close()

        # while fileisopen(global_element.reportpath):
        #     user_choose = win32api.MessageBox(0, 'The report is opened, pls close it before click OK buttom!', 'Tip',
        #                                       win32con.MB_OK)
        #     time.sleep(1)

        rpwb = load_workbook(global_elements.REPORTPATH)

        sheetname_list = rpwb.sheetnames
        if 'DUT Info' not in sheetname_list:
            rpwb.create_sheet('DUT Info')
            columns_list = ['DUT Name:', 'DUT Brand:', 'DUT Hard Revision:', 'DUT Software Revision:',
                            'DUT SN.:', 'DUT IMEI:', 'Maximum Registration Time(s):', 'Voltage High(V):',
                            'Voltage Normal(V):', 'Voltage Low(V):', 'Current Max(A):', 'Temperature High(℃):',
                            'Temperature Normal(℃):', 'Temperature Low(℃):']
            dut_config_dict = global_elements.dutActiveDict['xml']['DUTCONFIG']
            active_dutname = ''
            for k, v in global_elements.dutAcitveStateDict.items():
                if v:
                    active_dutname = k
                    break

            DUT_value_list = [active_dutname, dut_config_dict['Brand'], dut_config_dict['HWVersion'],
                              dut_config_dict['SWVersion'], dut_config_dict['SN'], dut_config_dict['IMEI'],
                              dut_config_dict['MAXREGTIME'], dut_config_dict['HV'], dut_config_dict['NV'],
                              dut_config_dict['LV'], dut_config_dict['MAXC'], dut_config_dict['HT'],
                              dut_config_dict['NT'],
                              dut_config_dict['LT']]
            sheet = rpwb['DUT Info']
            for i in range(len(columns_list)):
                sheet.cell(row=i + 1, column=1).value = columns_list[i]
                sheet.cell(row=i + 1, column=2).value = DUT_value_list[i]

        sheetname_list = rpwb.sheetnames
        if (' '.join(global_elements.test_case_name.split(' ')[:2])) not in sheetname_list:
            rpwb.create_sheet(' '.join(global_elements.test_case_name.split(' ')[:2]))
            list_head = ['Test Type', 'Test Case', 'Test Item', 'Result', 'Low Limit', 'High Limit', 'Unit',
                         'Judgement', 'Waveform', 'Test Band', 'BandWidth',
                         'SCS', 'Test ID', 'RB Allocation', 'MCS Index', 'Modulation', 'DL_Channel', 'UL_Channel',
                         'DL_Freq(MHz)', 'UL_Freq(MHz)', 'Volt.(V)', 'Temp.(℃)',
                         'Time', 'Remark']
            sheet = rpwb[' '.join(global_elements.test_case_name.split(' ')[:2])]
            for i in range(len(list_head)):
                sheet.cell(row=1, column=i + 1).value = list_head[i]

        sheet_case = rpwb[' '.join(global_elements.test_case_name.split(' ')[:2])]
        Judgement = judgement_handle(result, lowlimit, highlimit)
        # 如果是TX项，只出UL的信息，如果是RX项，只出DL的信息
        if 'ts138521_1 6.' in global_elements.test_case_name:
            UL_Freq = NreftoFre(int(global_elements.NR_ch_r))
            UL_channel = global_elements.NR_ch_r
            DL_Freq = '--'
            DL_channel = '--'
        else:
            UL_Freq = NreftoFre(int(global_elements.NR_ch_r))
            UL_channel = global_elements.NR_ch_r
            DL_Freq = NreftoFre(int(global_elements.NR_dl_ch_r))
            DL_channel = global_elements.NR_dl_ch_r

        time_now = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
        row_count = sheet_case.max_row
        list_data = [global_elements.NR_type_r, global_elements.test_case_name_r, global_elements.test_item_r, result,
                     lowlimit,
                     highlimit, unit, Judgement, global_elements.NR_Waveform_r, global_elements.NR_band_r,
                     global_elements.NR_bw_r, global_elements.NR_SCS_r, global_elements.NR_testid_r,
                     global_elements.NR_RBAllocation_r,
                     global_elements.NR_MCSindex_r, global_elements.NR_Modulation_r, DL_channel, UL_channel,
                     DL_Freq, UL_Freq,
                     global_elements.dutActiveDict['xml']['DUTCONFIG'][global_elements.NR_condition[2:4]],
                     global_elements.dutActiveDict['xml']['DUTCONFIG'][global_elements.NR_condition[0:2]], time_now,
                     global_elements.NR_remark_r]
        for i in range(len(list_data)):
            sheet_case.cell(row=(row_count + 1), column=i + 1).value = list_data[i]
            if i == 7:
                if Judgement == 'Passed':
                    sheet_case.cell(row=(row_count + 1), column=i + 1).font = font_pass
                elif Judgement == 'Failed':
                    sheet_case.cell(row=(row_count + 1), column=i + 1).font = font_fail
                else:
                    sheet_case.cell(row=(row_count + 1), column=i + 1).font = font_inconclusive

        for i in sheetname_list:
            if 'Sheet' in i:
                rpwb.remove(rpwb[i])
                break
        rpwb.save(global_elements.REPORTPATH)

        global_elements.emitsingle.reportupdataSingle.emit(list_data)  # 更新log界面
        if Judgement != 'Passed':
            global_elements.emitsingle.faileditemsupdataSingle.emit(list_data)  # 如果结果不是Pass,将结果更新到failed items界面

        if Judgement == 'Passed':
            global_elements.finished_result_list[0] += 1
        elif Judgement == 'Failed':
            global_elements.finished_result_list[1] += 1
        elif Judgement == 'Inconclusive':
            global_elements.finished_result_list[2] += 1


def ReporttoolforNSA(result, lowlimit, highlimit, unit):
    if global_elements.isStatusError == False:  # 判断是否已经终止进程
        if not os.access(global_elements.REPORTPATH, os.F_OK):  # 先判断report文件是否存存在， 不存在就新建
            reportwb = Workbook()
            sheets_list = reportwb.sheetnames
            if len(sheets_list) > 1:
                for i in range(1, len(sheets_list)):
                    reportwb.remove(reportwb[sheets_list[i]])  # 只保存一个sheet

            reportwb.save(global_elements.REPORTPATH)
            reportwb.close()

        # while fileisopen(global_element.reportpath):
        #     user_choose = win32api.MessageBox(0, 'The report is opened, pls close it before click OK buttom!', 'Tip',
        #                                       win32con.MB_OK)
        #     time.sleep(1)

        rpwb = load_workbook(global_elements.REPORTPATH)

        sheetname_list = rpwb.sheetnames
        if 'DUT Info' not in sheetname_list:
            rpwb.create_sheet('DUT Info')
            columns_list = ['DUT Name:', 'DUT Brand:', 'DUT Hard Revision:', 'DUT Software Revision:',
                            'DUT SN.:', 'DUT IMEI:', 'Maximum Registration Time(s):', 'Voltage High(V):',
                            'Voltage Normal(V):', 'Voltage Low(V):', 'Current Max(A):', 'Temperature High(℃):',
                            'Temperature Normal(℃):', 'Temperature Low(℃):']
            dut_config_dict = global_elements.dutActiveDict['xml']['DUTCONFIG']
            active_dutname = ''
            for k, v in global_elements.dutAcitveStateDict.items():
                if v:
                    active_dutname = k
                    break

            DUT_value_list = [active_dutname, dut_config_dict['Brand'], dut_config_dict['HWVersion'],
                              dut_config_dict['SWVersion'], dut_config_dict['SN'], dut_config_dict['IMEI'],
                              dut_config_dict['MAXREGTIME'], dut_config_dict['HV'], dut_config_dict['NV'],
                              dut_config_dict['LV'], dut_config_dict['MAXC'], dut_config_dict['HT'],
                              dut_config_dict['NT'],
                              dut_config_dict['LT']]
            sheet = rpwb['DUT Info']
            for i in range(len(columns_list)):
                sheet.cell(row=i + 1, column=1).value = columns_list[i]
                sheet.cell(row=i + 1, column=2).value = DUT_value_list[i]

        sheetname_list = rpwb.sheetnames
        if (' '.join(global_elements.test_case_name.split(' ')[:2])) not in sheetname_list:
            rpwb.create_sheet(' '.join(global_elements.test_case_name.split(' ')[:2]))
            list_head = ['Test Type', 'Test Case', 'Test Item', 'Result', 'Low Limit', 'High Limit', 'Unit',
                         'Judgement', 'Waveform', 'Test Band', 'BandWidth',
                         'SCS', 'Test ID', 'RB Allocation', 'MCS Index', 'Modulation', 'DL_Channel', 'UL_Channel',
                         'DL_Freq(MHz)', 'UL_Freq(MHz)', 'Volt.(V)', 'Temp.(℃)',
                         'Time', 'Remark']
            sheet = rpwb[' '.join(global_elements.test_case_name.split(' ')[:2])]
            for i in range(len(list_head)):
                sheet.cell(row=1, column=i + 1).value = list_head[i]

        sheet_case = rpwb[' '.join(global_elements.test_case_name.split(' ')[:2])]
        Judgement = judgement_handle(result, lowlimit, highlimit)
        # 如果是TX项，只出UL的信息，如果是RX项，只出DL的信息
        nr_UL_Freq = global_elements.NSA_NR_ULFreq
        nr_UL_channel = global_elements.NSA_NR_channel
        lte_UL_Freq = global_elements.NSA_LTE_ULFreq
        lte_UL_channel = global_elements.NSA_LTE_ULchannel
        UL_Freq = 'LTE:' + lte_UL_Freq + ' & NR:' + nr_UL_Freq
        UL_channel = 'LTE:' + lte_UL_channel + ' & NR:' + nr_UL_channel
        if 'ts138521_3 6.' in global_elements.test_case_name:
            DL_Freq = '--'
            DL_channel = '--'
        else:
            nr_DL_Freq = global_elements.NSA_NR_DLFreq
            nr_DL_channel = global_elements.NSA_NR_DLchannel
            lte_DL_Freq = global_elements.NSA_LTE_DLFreq
            lte_DL_channel = global_elements.NSA_LTE_DLchannel
            DL_Freq = 'LTE:' + lte_DL_Freq + ' & NR:' + nr_DL_Freq
            DL_channel = 'LTE:' + lte_DL_channel + ' & NR:' + nr_DL_channel

        bw_str_r = 'LTE:' + global_elements.NSA_LTE_bw_r + ' & NR:' + global_elements.NSA_NR_bw_r
        rb_str_r = 'LTE:' + global_elements.NSA_LTE_rb_r + ' & NR:' + global_elements.NSA_NR_rb_r
        modulation_str_r = 'LTE:' + global_elements.NSA_LTE_modulation + ' & NR:' + global_elements.NSA_NR_modulation
        time_now = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
        row_count = sheet_case.max_row
        list_data = [global_elements.NR_type_r, global_elements.test_case_name_r, global_elements.test_item_r, result,
                     lowlimit,
                     highlimit, unit, Judgement, global_elements.NR_Waveform_r, global_elements.NR_band_r,
                     bw_str_r, global_elements.NR_SCS_r, global_elements.NR_testid_r, rb_str_r,
                     global_elements.NR_MCSindex_r, modulation_str_r, DL_channel, UL_channel,
                     DL_Freq, UL_Freq,
                     global_elements.dutActiveDict['xml']['DUTCONFIG'][global_elements.NR_condition[2:4]],
                     global_elements.dutActiveDict['xml']['DUTCONFIG'][global_elements.NR_condition[0:2]], time_now,
                     global_elements.NR_remark_r]
        for i in range(len(list_data)):
            sheet_case.cell(row=(row_count + 1), column=i + 1).value = list_data[i]
            if i == 7:
                if Judgement == 'Passed':
                    sheet_case.cell(row=(row_count + 1), column=i + 1).font = font_pass
                elif Judgement == 'Failed':
                    sheet_case.cell(row=(row_count + 1), column=i + 1).font = font_fail
                else:
                    sheet_case.cell(row=(row_count + 1), column=i + 1).font = font_inconclusive

        for i in sheetname_list:
            if 'Sheet' in i:
                rpwb.remove(rpwb[i])
                break
        rpwb.save(global_elements.REPORTPATH)

        global_elements.emitsingle.reportupdataSingle.emit(list_data)  # 更新log界面
        if Judgement != 'Passed':
            global_elements.emitsingle.faileditemsupdataSingle.emit(list_data)  # 如果结果不是Pass,将结果更新到failed items界面

        if Judgement == 'Passed':
            global_elements.finished_result_list[0] += 1
        elif Judgement == 'Failed':
            global_elements.finished_result_list[1] += 1
        elif Judgement == 'Inconclusive':
            global_elements.finished_result_list[2] += 1
